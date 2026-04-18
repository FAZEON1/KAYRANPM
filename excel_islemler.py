import pandas as pd
from datetime import datetime
from database import upsert_urun, upsert_firma_stok, get_connection, upsert_yoldaki_urun

FIRMA_LISTESI = ["ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DIGER"]

def excel_yukle_ana_stok(dosya_yolu):
    """
    Ana stok sekmesini yükler.
    Beklenen kolonlar: SKU, Ürün Adı, Kategori, Marka, Fiyat, Özellikler, Bizim Stok, Trendyol Stok
    """
    try:
        df = pd.read_excel(dosya_yolu, sheet_name=0)  # İlk sekmeyi oku (G5F STOK)
        df.columns = [str(c).strip().upper() for c in df.columns]
        
        kolon_esleme = {
            "SKU": ["SKU", "KOD", "ÜRÜN KODU", "URUN KODU", "BARKOD"],
            "URUN_ADI": ["ÜRÜN ADI", "URUN ADI", "AD", "ÜRÜN", "URUN", "PRODUCT"],
            "KATEGORI": ["KATEGORİ", "KATEGORI", "CATEGORY"],
            "MARKA": ["MARKA", "BRAND"],
            "SATIS_FIYATI": ["SATIŞ FİYATI", "SATIS FIYATI", "FİYAT", "FIYAT", "PRICE", "SATIŞ FİYATI"],
            "ALIS_FIYATI": ["ALIŞ FİYATI", "ALIS FIYATI", "MALİYET", "MALIYET", "COST"],
            "HEDEF_KAR": ["HEDEF KAR MARJI", "HEDEF KAR", "KAR MARJI", "MARGIN"],
            "OZELLIKLER": ["ÖZELLİKLER", "OZELLIKLER", "SPEC", "AÇIKLAMA"],
            "BIZIM_STOK": ["BİZİM STOK", "BIZIM STOK", "DEPO STOK", "STOK", "G5F STOK"],
            "TRENDYOL_STOK": ["TRENDYOL STOK", "TRENDYOL", "TY STOK"],
            "YOLDAKI_MIKTAR": ["YOLDAKİ MİKTAR", "YOLDAKI MIKTAR", "YOL MİKTAR", "YOLDA"],
            "VARIS_TARIHI": ["TAHMİNİ VARIŞ", "TAHMINI VARIS", "VARIŞ TARİHİ", "VARIS TARIHI", "ETA"],
        }
        
        kolon_map = {}
        for hedef, alternatifler in kolon_esleme.items():
            for alt in alternatifler:
                if alt in df.columns:
                    kolon_map[hedef] = alt
                    break
        
        if "SKU" not in kolon_map:
            return False, "SKU/Ürün Kodu kolonu bulunamadı."
        if "URUN_ADI" not in kolon_map:
            return False, "Ürün Adı kolonu bulunamadı."
        
        basarili = 0
        hatali = 0
        for _, row in df.iterrows():
            try:
                sku = str(row[kolon_map["SKU"]]).strip()
                if not sku or sku == "nan":
                    continue
                urun_adi = str(row.get(kolon_map.get("URUN_ADI", ""), "")).strip()
                kategori = str(row.get(kolon_map.get("KATEGORI", ""), "")).strip() if "KATEGORI" in kolon_map else ""
                marka = str(row.get(kolon_map.get("MARKA", ""), "")).strip() if "MARKA" in kolon_map else ""
                ozellikler = str(row.get(kolon_map.get("OZELLIKLER", ""), "")).strip() if "OZELLIKLER" in kolon_map else ""
                try: satis_fiyati = float(row.get(kolon_map.get("SATIS_FIYATI", ""), 0) or 0)
                except: satis_fiyati = 0.0
                try: alis_fiyati = float(row.get(kolon_map.get("ALIS_FIYATI", ""), 0) or 0)
                except: alis_fiyati = 0.0
                try: hedef_kar = float(row.get(kolon_map.get("HEDEF_KAR", ""), 0) or 0)
                except: hedef_kar = 0.0
                try: bizim_stok = int(row.get(kolon_map.get("BIZIM_STOK", ""), 0) or 0)
                except: bizim_stok = 0
                try: trendyol_stok = int(row.get(kolon_map.get("TRENDYOL_STOK", ""), 0) or 0)
                except: trendyol_stok = 0

                # Yoldaki bilgisi G5F STOK sekmesinden oku
                try: yoldaki_miktar = int(row.get(kolon_map.get("YOLDAKI_MIKTAR", ""), 0) or 0)
                except: yoldaki_miktar = 0
                try:
                    varis_tarihi = str(row.get(kolon_map.get("VARIS_TARIHI", ""), "") or "").strip()
                    if varis_tarihi == "nan": varis_tarihi = ""
                except: varis_tarihi = ""

                upsert_urun(sku, urun_adi, kategori, marka, satis_fiyati, alis_fiyati, hedef_kar, ozellikler, bizim_stok, trendyol_stok)

                # Yoldaki veriyi de kaydet
                if yoldaki_miktar > 0 or varis_tarihi:
                    upsert_yoldaki_urun(sku, urun_adi, yoldaki_miktar, varis_tarihi)
                basarili += 1
            except Exception as e:
                hatali += 1
        
        return True, f"{basarili} ürün başarıyla yüklendi. {hatali} satır atlandı."
    except Exception as e:
        return False, f"Dosya okunamadı: {str(e)}"


def excel_yukle_firma_stoklari(dosya_yolu):
    """
    Firma stok sekmelerini yükler.
    Her firma için ayrı sekme: ITOPYA, HB, VATAN, MONDAY, KANAL, DIGER
    Beklenen kolonlar: SKU, Ürün Adı, Stok Miktarı, Haftalık Satış
    """
    try:
        xl = pd.ExcelFile(dosya_yolu)
        mevcut_sekmeler = [s.strip().upper() for s in xl.sheet_names]
        
        sonuclar = []
        for firma in FIRMA_LISTESI:
            # Sekme adı eşleştirme (DİĞER -> DIGER vb.)
            eslesen_sekme = None
            for sekme in xl.sheet_names:
                if sekme.strip().upper().replace("İ", "I").replace("Ğ", "G").replace("Ü", "U").replace("Ş", "S").replace("Ç", "C").replace("Ö", "O") == firma.replace("İ", "I").replace("Ğ", "G").replace("Ü", "U").replace("Ş", "S").replace("Ç", "C").replace("Ö", "O"):
                    eslesen_sekme = sekme
                    break
                if firma in sekme.strip().upper():
                    eslesen_sekme = sekme
                    break
            
            if not eslesen_sekme:
                sonuclar.append(f"⚠️ {firma}: Sekme bulunamadı, atlandı.")
                continue
            
            df = pd.read_excel(dosya_yolu, sheet_name=eslesen_sekme)
            df.columns = [str(c).strip().upper() for c in df.columns]
            
            kolon_esleme = {
                "SKU": ["SKU", "KOD", "ÜRÜN KODU", "URUN KODU", "BARKOD"],
                "URUN_ADI": ["ÜRÜN ADI", "URUN ADI", "AD", "ÜRÜN", "URUN"],
                "STOK": ["STOK MİKTARI", "STOK MIKTARI", "STOK", "MEVCUT STOK", "ADET"],
                "SATIS": ["HAFTALIK SATIŞ", "HAFTALIK SATIS", "SATIŞ", "SATIS", "SATIS ADEDI"],
            }
            
            kolon_map = {}
            for hedef, alternatifler in kolon_esleme.items():
                for alt in alternatifler:
                    if alt in df.columns:
                        kolon_map[hedef] = alt
                        break
            
            if "SKU" not in kolon_map:
                sonuclar.append(f"❌ {firma}: SKU kolonu bulunamadı.")
                continue
            
            basarili = 0
            for _, row in df.iterrows():
                try:
                    sku = str(row[kolon_map["SKU"]]).strip()
                    if not sku or sku == "nan":
                        continue
                    urun_adi = str(row.get(kolon_map.get("URUN_ADI", ""), "")).strip() if "URUN_ADI" in kolon_map else ""
                    stok = int(row.get(kolon_map.get("STOK", ""), 0) or 0) if "STOK" in kolon_map else 0
                    satis = int(row.get(kolon_map.get("SATIS", ""), 0) or 0) if "SATIS" in kolon_map else 0
                    upsert_firma_stok(firma, sku, urun_adi, stok, satis)
                    basarili += 1
                except:
                    pass
            
            sonuclar.append(f"✅ {firma}: {basarili} ürün yüklendi.")
        
        return True, "\n".join(sonuclar)
    except Exception as e:
        return False, f"Dosya okunamadı: {str(e)}"


def excel_yukle_yoldaki_urunler(dosya_yolu):
    """
    Yoldaki ürünler sekmesini yükler.
    Beklenen kolonlar: SKU, Ürün Adı, Yoldaki Miktar, Tahmini Varış Tarihi
    """
    try:
        xl = pd.ExcelFile(dosya_yolu)
        eslesen_sekme = None
        for sekme in xl.sheet_names:
            s = sekme.strip().upper()
            if "YOLDAK" in s or "YOL" in s or "TRANSIT" in s or "SIPARIS" in s:
                eslesen_sekme = sekme
                break

        if not eslesen_sekme:
            return False, "❌ 'YOLDAKI' adında sekme bulunamadı."

        df = pd.read_excel(dosya_yolu, sheet_name=eslesen_sekme)
        df.columns = [str(c).strip().upper() for c in df.columns]

        kolon_esleme = {
            "SKU": ["SKU", "KOD", "ÜRÜN KODU", "URUN KODU"],
            "URUN_ADI": ["ÜRÜN ADI", "URUN ADI", "AD", "ÜRÜN"],
            "MIKTAR": ["YOLDAKI MIKTAR", "YOLDAKİ MİKTAR", "MİKTAR", "MIKTAR", "ADET", "SİPARİŞ MİKTARI"],
            "VARIS": ["TAHMİNİ VARIŞ", "TAHMINI VARIS", "VARIŞ TARİHİ", "VARIS TARIHI", "ETD", "ETA"],
        }

        kolon_map = {}
        for hedef, alternatifler in kolon_esleme.items():
            for alt in alternatifler:
                if alt in df.columns:
                    kolon_map[hedef] = alt
                    break

        if "SKU" not in kolon_map:
            return False, "❌ SKU kolonu bulunamadı."

        basarili = 0
        for _, row in df.iterrows():
            try:
                sku = str(row[kolon_map["SKU"]]).strip()
                if not sku or sku == "nan":
                    continue
                urun_adi = str(row.get(kolon_map.get("URUN_ADI", ""), "")).strip() if "URUN_ADI" in kolon_map else ""
                miktar = int(row.get(kolon_map.get("MIKTAR", ""), 0) or 0) if "MIKTAR" in kolon_map else 0
                varis = str(row.get(kolon_map.get("VARIS", ""), "")).strip() if "VARIS" in kolon_map else ""
                if varis == "nan":
                    varis = ""
                upsert_yoldaki_urun(sku, urun_adi, miktar, varis)
                basarili += 1
            except:
                pass

        return True, f"✅ Yoldaki ürünler: {basarili} satır yüklendi."
    except Exception as e:
        return False, f"Dosya okunamadı: {str(e)}"



def create_sample_excel_bytes():
    """Örnek Excel şablonunu bellekte oluşturur ve bytes döndürür (Streamlit Cloud için)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "G5F STOK"
    basliklar = ["SKU", "Ürün Adı", "Kategori", "Marka", "Satış Fiyatı (₺)", "Alış Fiyatı (₺)", "Hedef Kar Marjı (%)", "Özellikler", "Bizim Stok", "Trendyol Stok", "Yoldaki Miktar", "Tahmini Varış Tarihi"]
    for i, b in enumerate(basliklar, 1):
        cell = ws1.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    for row in [
        ["SKU001", "Samsung Galaxy S24", "Telefon", "Samsung", 35000, 22000, 30, "6.2 inch, 128GB", 50, 30, 100, "2026-05-15"],
        ["SKU002", "iPhone 15", "Telefon", "Apple", 55000, 38000, 25, "6.1 inch, 128GB", 20, 15, 0, ""],
        ["SKU003", "Xiaomi Redmi Note 13", "Telefon", "Xiaomi", 12000, 7500, 35, "6.67 inch, 256GB", 80, 60, 50, "2026-06-01"],
    ]:
        ws1.append(row)
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18

    for firma in ["ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DIGER"]:
        ws = wb.create_sheet(firma)
        for i, b in enumerate(["SKU", "Ürün Adı", "Stok Miktarı", "Haftalık Satış"], 1):
            cell = ws.cell(row=1, column=i, value=b)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="2E7D32")
            cell.alignment = Alignment(horizontal="center")
        for row in [["SKU001", "Samsung Galaxy S24", 10, 5], ["SKU002", "iPhone 15", 3, 2], ["SKU003", "Xiaomi Redmi Note 13", 25, 12]]:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    ws_yol = wb.create_sheet("YOLDAKI")
    for i, b in enumerate(["SKU", "Ürün Adı", "Yoldaki Miktar", "Tahmini Varış Tarihi"], 1):
        cell = ws_yol.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6A1B9A")
        cell.alignment = Alignment(horizontal="center")
    for row in [["SKU001", "Samsung Galaxy S24", 30, "2026-05-01"], ["SKU002", "iPhone 15", 10, "2026-04-25"]]:
        ws_yol.append(row)
    for col in ws_yol.columns:
        ws_yol.column_dimensions[col[0].column_letter].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def create_sample_excel():
    """Örnek Excel şablonu oluşturur"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Ana Stok sekmesi
    ws1 = wb.active
    ws1.title = "G5F STOK"
    basliklar = ["SKU", "Ürün Adı", "Kategori", "Marka", "Satış Fiyatı (₺)", "Alış Fiyatı (₺)", "Hedef Kar Marjı (%)", "Özellikler", "Bizim Stok", "Trendyol Stok", "Yoldaki Miktar", "Tahmini Varış Tarihi"]
    for i, b in enumerate(basliklar, 1):
        cell = ws1.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    ornek_veri = [
        ["SKU001", "Samsung Galaxy S24", "Telefon", "Samsung", 35000, 22000, 30, "6.2 inch, 128GB", 50, 30, 100, "2026-05-15"],
        ["SKU002", "iPhone 15", "Telefon", "Apple", 55000, 38000, 25, "6.1 inch, 128GB", 20, 15, 0, ""],
        ["SKU003", "Xiaomi Redmi Note 13", "Telefon", "Xiaomi", 12000, 7500, 35, "6.67 inch, 256GB", 80, 60, 50, "2026-06-01"],
    ]
    for row in ornek_veri:
        ws1.append(row)
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 18
    
    # Firma sekmeleri
    firma_listesi_tr = ["ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DIGER"]
    for firma in firma_listesi_tr:
        ws = wb.create_sheet(firma)
        firma_basliklar = ["SKU", "Ürün Adı", "Stok Miktarı", "Haftalık Satış"]
        for i, b in enumerate(firma_basliklar, 1):
            cell = ws.cell(row=1, column=i, value=b)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", start_color="2E7D32")
            cell.alignment = Alignment(horizontal="center")
        ornek = [
            ["SKU001", "Samsung Galaxy S24", 10, 5],
            ["SKU002", "iPhone 15", 3, 2],
            ["SKU003", "Xiaomi Redmi Note 13", 25, 12],
        ]
        for row in ornek:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

    # Yoldaki ürünler sekmesi
    ws_yol = wb.create_sheet("YOLDAKI")
    yol_basliklar = ["SKU", "Ürün Adı", "Yoldaki Miktar", "Tahmini Varış Tarihi"]
    for i, b in enumerate(yol_basliklar, 1):
        cell = ws_yol.cell(row=1, column=i, value=b)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="6A1B9A")
        cell.alignment = Alignment(horizontal="center")
    ornek_yol = [
        ["SKU001", "Samsung Galaxy S24", 30, "2026-05-01"],
        ["SKU002", "iPhone 15", 10, "2026-04-25"],
    ]
    for row in ornek_yol:
        ws_yol.append(row)
    for col in ws_yol.columns:
        ws_yol.column_dimensions[col[0].column_letter].width = 22

    path = "/home/claude/stok_app/SABLON_STOK_TAKIP.xlsx"
    wb.save(path)
    return path
