from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from datetime import datetime
from analitik import dashboard_hesapla
from database import get_siparis_onerileri

RENKLER = {
    "kirmizi": "FFCCCC",
    "turuncu": "FFD580",
    "sari": "FFFF99",
    "yesil": "CCFFCC",
    "yok": "FFFFFF",
}

PERFORMANS_RENKLER = {
    "Çok İyi": "CCFFCC",
    "İyi": "FFFACD",
    "Düşük": "FFCCCC",
    "veri yok": "F0F0F0",
}

def thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def excel_rapor_olustur(kayit_yolu):
    try:
        data = dashboard_hesapla()
        wb = Workbook()
        
        # ---- SHEET 1: DASHBOARD ----
        ws = wb.active
        ws.title = "Dashboard"
        
        baslik_font = Font(bold=True, color="FFFFFF", size=10)
        baslik_fill = PatternFill("solid", start_color="1F4E79")
        merkez = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        basliklar = [
            "SKU", "Ürün Adı", "Kategori", "Marka",
            "Bizim Stok", "Trendyol Stok", "Stok Yaşı (Gün)",
            "Firma", "Firma Stok", "Haftalık Satış",
            "Kaç Günlük Satış", "Performans", "Sipariş Uyarısı"
        ]
        
        for i, b in enumerate(basliklar, 1):
            cell = ws.cell(row=1, column=i, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = merkez
            cell.border = thin_border()
        ws.row_dimensions[1].height = 35
        ws.freeze_panes = "A2"
        
        satir = 2
        for urun in data:
            for fd in urun["firma_detay"]:
                stok_fill = PatternFill("solid", start_color=RENKLER.get(urun["stok_renk"], "FFFFFF"))
                gun_fill = PatternFill("solid", start_color=RENKLER.get(fd["gun_renk"], "FFFFFF"))
                perf_fill = PatternFill("solid", start_color=PERFORMANS_RENKLER.get(fd["performans"], "FFFFFF"))
                
                degerler = [
                    urun["sku"],
                    urun["urun_adi"],
                    urun["kategori"],
                    urun["marka"],
                    urun["bizim_stok"],
                    urun["trendyol_stok"],
                    urun["stok_gun"],
                    fd["firma"],
                    fd["stok"],
                    fd["satis"],
                    fd["gun_sayisi"] if fd["gun_sayisi"] is not None else "-",
                    fd["performans"],
                    "⚠️ SİPARİŞ ÖNERİ!" if fd["siparis_uyarisi"] else "",
                ]
                
                for j, val in enumerate(degerler, 1):
                    cell = ws.cell(row=satir, column=j, value=val)
                    cell.border = thin_border()
                    cell.alignment = Alignment(vertical="center")
                    
                    if j == 7:  # Stok yaşı
                        cell.fill = stok_fill
                    elif j == 11:  # Kaç günlük satış
                        cell.fill = gun_fill
                    elif j == 12:  # Performans
                        cell.fill = perf_fill
                    elif j == 13 and fd["siparis_uyarisi"]:  # Uyarı
                        cell.fill = PatternFill("solid", start_color="FF0000")
                        cell.font = Font(bold=True, color="FFFFFF")
                
                satir += 1
        
        genislikler = [12, 30, 15, 15, 12, 14, 14, 12, 12, 14, 15, 12, 18]
        for i, g in enumerate(genislikler, 1):
            ws.column_dimensions[get_column_letter(i)].width = g
        
        # ---- SHEET 2: STOK YAYILIMI ----
        ws2 = wb.create_sheet("Stok Yayılımı")
        yayilim_baslik = ["SKU", "Ürün Adı", "Bizim Stok", "TRENDYOL", "ITOPYA", "HB", "VATAN", "MONDAY", "KANAL", "DİĞER", "Toplam Kanal Stok"]
        for i, b in enumerate(yayilim_baslik, 1):
            cell = ws2.cell(row=1, column=i, value=b)
            cell.font = baslik_font
            cell.fill = baslik_fill
            cell.alignment = merkez
            cell.border = thin_border()
        
        for satir_no, urun in enumerate(data, 2):
            y = urun["yayilim"]
            toplam = sum(y.values())
            row_data = [
                urun["sku"], urun["urun_adi"], urun["bizim_stok"],
                y.get("TRENDYOL", 0), y.get("ITOPYA", 0), y.get("HB", 0),
                y.get("VATAN", 0), y.get("MONDAY", 0), y.get("KANAL", 0),
                y.get("DIGER", 0), toplam
            ]
            for j, val in enumerate(row_data, 1):
                cell = ws2.cell(row=satir_no, column=j, value=val)
                cell.border = thin_border()
        
        for i, g in enumerate([12, 30, 12, 12, 12, 10, 12, 12, 12, 10, 16], 1):
            ws2.column_dimensions[get_column_letter(i)].width = g
        
        # ---- SHEET 3: SİPARİŞ ÖNERİLERİ ----
        ws3 = wb.create_sheet("Sipariş Önerileri")
        sp_baslik = ["ID", "Firma", "SKU", "Ürün Adı", "Önerilen Miktar", "Durum", "Oluşturma Tarihi", "Onay Tarihi"]
        for i, b in enumerate(sp_baslik, 1):
            cell = ws3.cell(row=1, column=i, value=b)
            cell.font = baslik_font
            cell.fill = PatternFill("solid", start_color="1B5E20")
            cell.alignment = merkez
            cell.border = thin_border()
        
        onerileri = get_siparis_onerileri()
        for satir_no, sp in enumerate(onerileri, 2):
            row_data = [sp["id"], sp["firma"], sp["sku"], sp["urun_adi"],
                        sp["oneri_miktari"], sp["durum"], sp["olusturma_tarihi"], sp.get("onay_tarihi", "")]
            for j, val in enumerate(row_data, 1):
                cell = ws3.cell(row=satir_no, column=j, value=val)
                cell.border = thin_border()
                if sp["durum"] == "onaylandi":
                    cell.fill = PatternFill("solid", start_color="CCFFCC")
                elif sp["durum"] == "reddedildi":
                    cell.fill = PatternFill("solid", start_color="FFCCCC")
        
        wb.save(kayit_yolu)
        return True, f"Excel raporu oluşturuldu: {kayit_yolu}"
    except Exception as e:
        return False, f"Excel raporu oluşturulamadı: {str(e)}"


def pdf_rapor_olustur(kayit_yolu):
    try:
        data = dashboard_hesapla()
        doc = SimpleDocTemplate(
            kayit_yolu,
            pagesize=landscape(A4),
            rightMargin=1*cm, leftMargin=1*cm,
            topMargin=1.5*cm, bottomMargin=1*cm
        )
        
        styles = getSampleStyleSheet()
        baslik_style = ParagraphStyle('baslik', parent=styles['Title'], fontSize=16, spaceAfter=12, textColor=colors.HexColor('#1F4E79'))
        alt_baslik_style = ParagraphStyle('altbaslik', parent=styles['Heading2'], fontSize=11, spaceAfter=6, textColor=colors.HexColor('#2E7D32'))
        
        story = []
        story.append(Paragraph("Ürün Yönetimi - Stok Takip Raporu", baslik_style))
        story.append(Paragraph(f"Oluşturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 0.5*cm))
        
        # Ana tablo başlıkları
        tablo_baslik = ["SKU", "Ürün Adı", "Biz.\nStok", "Stok\nYaşı", "Firma", "Firma\nStok", "Haftalık\nSatış", "Kaç Gün", "Performans", "Uyarı"]
        tablo_veri = [tablo_baslik]
        
        PDF_RENKLER = {
            "kirmizi": colors.HexColor('#FFCCCC'),
            "turuncu": colors.HexColor('#FFD580'),
            "sari": colors.HexColor('#FFFF99'),
            "yesil": colors.HexColor('#CCFFCC'),
            "yok": colors.white,
        }
        
        satir_stilleri = []
        satir_no = 1
        
        for urun in data:
            for fd in urun["firma_detay"]:
                satir = [
                    urun["sku"],
                    urun["urun_adi"][:25] + ("..." if len(urun["urun_adi"]) > 25 else ""),
                    str(urun["bizim_stok"]),
                    f"{urun['stok_gun']}g",
                    fd["firma"],
                    str(fd["stok"]),
                    str(fd["satis"]),
                    f"{fd['gun_sayisi']}g" if fd["gun_sayisi"] is not None else "-",
                    fd["performans"],
                    "⚠️" if fd["siparis_uyarisi"] else "",
                ]
                tablo_veri.append(satir)
                
                # Stok yaşı rengi (sütun 3)
                stok_renk = PDF_RENKLER.get(urun["stok_renk"], colors.white)
                satir_stilleri.append(('BACKGROUND', (3, satir_no), (3, satir_no), stok_renk))
                
                # Kaç günlük satış rengi (sütun 7)
                gun_renk = PDF_RENKLER.get(fd["gun_renk"], colors.white)
                satir_stilleri.append(('BACKGROUND', (7, satir_no), (7, satir_no), gun_renk))
                
                # Sipariş uyarısı
                if fd["siparis_uyarisi"]:
                    satir_stilleri.append(('BACKGROUND', (9, satir_no), (9, satir_no), colors.HexColor('#FF4444')))
                    satir_stilleri.append(('TEXTCOLOR', (9, satir_no), (9, satir_no), colors.white))
                
                satir_no += 1
        
        col_widths = [2.5*cm, 6*cm, 1.5*cm, 1.5*cm, 2*cm, 1.5*cm, 2*cm, 1.8*cm, 2.5*cm, 1.5*cm]
        
        tablo = Table(tablo_veri, colWidths=col_widths, repeatRows=1)
        tablo_stil = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ] + satir_stilleri)
        tablo.setStyle(tablo_stil)
        
        story.append(tablo)
        doc.build(story)
        return True, f"PDF raporu oluşturuldu: {kayit_yolu}"
    except Exception as e:
        return False, f"PDF raporu oluşturulamadı: {str(e)}"
